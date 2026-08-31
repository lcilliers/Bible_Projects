"""cfgquality.py — shared config-quality checks, used by BOTH handlers/configmaint.py (the
propose-time / validate-time checks) and lib/cfgreport.py (so CONFIG-REPORT.md can show current
findings, not just the live escalation). Split out 2026-07-21 to avoid a circular import
(configmaint.py already imports cfgreport; cfgreport importing configmaint back would cycle).
"""

from __future__ import annotations

import io
import json
import pathlib
import re
import sqlite3
import tokenize

# module -> the dedicated table it already has, if any. Per the researcher's 2026-07-21 rule
# ("there must be a very good reason why a config goes into settings, rather than the specific
# module or utility"): a module on this list already has a purpose-built home, so a NEW
# cfg_setting row for it needs explicit justification. Grows the same way CFG_TABLES does — a
# small, named fact, not derived generically (only one entry exists so far).
MODULE_DEDICATED_TABLE = {
    "candidate": "cfg_candidate_rule",
}

# module -> (table, key) locating the setting that must hold WHERE that module's quality-check
# findings persist. `table` is almost always 'cfg_setting', but a module with its own dedicated
# settings table (governance.module.config — e.g. cfg_passage, escalation #798/#799) keeps its
# report-path setting there instead; each entry names its own home rather than assuming one table
# for all. Per the researcher's 2026-07-21 rule ("errors is not optional to fix... why is there a
# standard if you don't follow it"): every step whose Outcome is advisory findings (not a data
# write) must persist those findings to a report file, matching report.py/validation.py/
# cfgreport.py's established pattern — not just a terminal print + an escalation row that scrolls
# away.
QUALITY_CHECK_REPORT_PATH = {
    "configmaint.validate": ("cfg_setting", "configmaint.report_path"),  # -> CONFIG-REPORT.md
    "candidate.validate": ("cfg_setting", "candidate.quality_report_path"),
    "passage.validate": ("cfg_passage", "passage.quality_report_path"),
    "lexicon.validate": ("cfg_setting", "lexicon.quality_report_path"),
}

# Every step known to write a persistent report via lib/reportkit.render_scaffold — the ground
# truth cfg_report SHOULD contain a row for, one per report-producing step. A hardcoded list, same
# shape as QUALITY_CHECK_REPORT_PATH above and for the same reason: checking cfg_report against
# itself couldn't catch a step that's missing its row entirely (added 2026-07-22,
# PLAN-reports-config-governance-v1-20260722.md §10.1/§11 — "will you miss configs when you build"
# gets a check, not a promise; this is the check that would have caught the retention/candidate.load
# gaps sooner had it existed then).
REPORT_STEPS = (
    "configmaint.report", "candidate.validate", "candidate.load", "passage.validate",
    "report.word", "validation.word", "validation.book", "retention.report",
    "report.seed_candidate", "report.strong_meaning", "report.span_analysis",
    "report.schema_overview", "report.registry", "lexicon.validate",
)


def _step_inactive(conn: sqlite3.Connection, step: str) -> bool:
    """True if `step` has an cfg_step row and every one of them is inactive — REPORT_STEPS/
    QUALITY_CHECK_REPORT_PATH are hardcoded Python names, disconnected from cfg_step.inactive
    (escalation #310), so checks keyed off them need to ask explicitly rather than silently keep
    flagging a step the researcher has already retired."""
    rows = conn.execute("SELECT inactive FROM cfg_step WHERE step=?", (step,)).fetchall()
    return bool(rows) and all(r[0] for r in rows)


def find_missing_cfg_report_rows(conn: sqlite3.Connection) -> list[str]:
    """Every ACTIVE step in REPORT_STEPS must have an active cfg_report row (title/ToC/footer/
    naming/CSV pairing) — a report-producing step with no cfg_report row means its content-shape
    is still hardcoded Python, the exact gap this plan closed for the original 8, re-checked so it
    can't silently reopen for a 9th. A retired (inactive) step is skipped entirely (escalation
    #310) — its missing/stale report config is no longer a live defect."""
    missing = []
    for step in REPORT_STEPS:
        if _step_inactive(conn, step):
            continue
        if not conn.execute(
                "SELECT 1 FROM cfg_report WHERE step=? AND inactive=0", (step,)).fetchone():
            missing.append(f"{step} produces a persistent report but has no active cfg_report row "
                          f"(title/sections/CSV pairing still hardcoded, or the row is inactive)")
    return missing


def find_chained_packages_missing_complete_message(conn: sqlite3.Connection) -> list[str]:
    """Every ACTIVE CHAINED work package prints a COMPLETE banner at the end of its sequence
    (PS-side) — if cfg_work_package.complete_message is NULL, that banner's wording has nowhere to
    come from but a hardcoded PS string again. An inactive work package is excluded (escalation
    #310) — a retired package with no complete_message isn't a live defect."""
    missing = []
    for r in conn.execute(
            "SELECT name FROM cfg_work_package WHERE chained=1 AND inactive=0 AND "
            "(complete_message IS NULL OR complete_message='')"):
        missing.append(f"{r[0]} is chained but has no cfg_work_package.complete_message")
    return missing


def find_settings_needing_justification(conn: sqlite3.Connection) -> list[str]:
    """A cfg_setting row whose module ALREADY has its own dedicated table — advisory. Doesn't
    (and can't) judge whether the reason is good — just surfaces every case where the question
    needs asking, so it isn't silently missed. Inactive settings are excluded (escalation #310) —
    a retired setting doesn't need a live justification decision."""
    flags = []
    for module, table in MODULE_DEDICATED_TABLE.items():
        for r in conn.execute(
                "SELECT key FROM cfg_setting WHERE module=? AND inactive=0", (module,)):
            flags.append(f"cfg_setting {r[0]!r} (module {module!r}) — {module} already has its "
                         f"own dedicated table ({table}); confirm this belongs in shared "
                         f"cfg_setting rather than there")
    return flags


# Modules whose cfg_setting rows are pure narrative/documentation -- a fact recorded for a human
# or an external, non-dispatcher-mediated process to read, never a value the app's own runtime
# applies via .setting(). Distinct from the module='governance' branch below: governance rows must
# still be read by init.py (AI-startup-compliance requirement); these have no such requirement at
# all. Escalation #719, 2026-08-18 -- start with 'backup' (verified: all 6 live rows document
# Windows Scheduled Tasks / standalone scripts outside the app, not one is a masked real orphan).
# Add a module here only after checking its live rows the same way, not by assumption.
_NARRATIVE_MODULES = frozenset({"backup"})

# module='database' keys of the shape `database.<name>.path` -- read via Cfg.database_path()'s
# `self.setting(f"database.{name}.path")`, an f-string-composed key the literal-string scan in
# find_orphan_configs cannot see. Narrowly matched against this exact known indirection (see the
# call-site check where this is used), not a blanket module exemption. Escalation #748.
_DATABASE_PATH_KEY = re.compile(r"^database\.[^.]+\.path$")


def find_orphan_configs(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """cfg_setting keys / cfg_enum groups without REAL usage — configs the app would not actually
    respond to if their value/membership changed. ADVISORY, not a coherence error: an orphan may
    be legitimately pre-staged for a not-yet-built step rather than a mistake.
    EXCLUDES iba/app/migration/: those scripts exist to WRITE a setting's initial value, so they
    always mention its name — counting that as "usage" would mask genuine orphans.

    REDEFINED 2026-07-23 (researcher's correction — escalation #305): "not referenced anywhere"
    was too loose a test — a key merely appearing as a quoted literal (e.g. in a comment or an
    unrelated docstring) passed it without the code actually applying the config's VALUE. "Usage"
    is not one shape; per the researcher, it differs by config kind:
      - a plain cfg_setting: the app must apply its VALUE at runtime. Proven by the key literal
        co-occurring, IN THE SAME FILE, with an actual `.setting(` accessor call — not just the
        key text appearing anywhere in the multi-file corpus, which could be satisfied by an
        unrelated comment/docstring in a file that never reads config at all. (Same-file rather
        than same-call-site: several settings are read via a level of indirection — e.g.
        validation.py's `_WORD_SECTIONS = {"label": "validation.show_health", ...}` then
        `cfg.setting(key, True)` in a loop — genuinely applied, just not through a literal
        `cfg.setting("validation.show_health", ...)` call site. Settings read via a
        cfg_column.expectation data-driven key, e.g. 'pattern:<key>', are handled by the
        exclusion below — also genuinely applied, also not through a literal call site.)
      - a cfg_setting with module='governance': these are process rules for the AI/researcher
        workflow, not runtime application inputs — there is no "apply the value" behaviour to grep
        for. Per the researcher: they "must be read by the startup routine explicitly to ensure
        that AI complies with it." Usage = referenced specifically in iba/app/init.py (the startup
        routine), not anywhere in the app at large — either by the individual key literal, or by a
        generic `WHERE module='governance'` read (init.py deliberately reads the whole module
        dynamically so a NEW governance setting is picked up without an init.py edit; that generic
        read counts as usage for every row it covers, the same reasoning as the
        cfg_column.expectation exclusion below).
      - a cfg_enum group: per the researcher, this is "a lookup, or options... not hard coded but
        use the config" — usage = the group is actually queried by NAME at runtime (`cfg.enum(name)`,
        or the equivalent raw `cfg_enum WHERE name='<name>'`/`name="<name>"` SQL some handlers use
        directly), so a change to the DB's membership is something code would notice. A group's
        VALUES appearing as hardcoded string literals elsewhere (e.g. `state == "paused"`) is NOT
        usage of the enum — the vocabulary isn't actually being read from cfg_enum in that case.

    FIXED 2026-07-22 (kept): a setting/enum read dynamically via `cfg_column.expectation` data
    ('pattern:<key>' for a cfg_setting, 'enum.<name>' for a cfg_enum group) is genuinely enforced
    by lib/valuequality.py's engine, but the value lives in a DB row, not literal .py source —
    excluded before the source-level checks below run."""
    # .ps1 too, not just .py: "PowerShell orchestrates, Python works" means a PS script reading
    # a setting via an inline `python -c "...c.setting('key')..."` (e.g. configmaint.auto_report,
    # read by Config-Maintenance.ps1 this way) is real usage the .py-only scan used to miss.
    per_file: dict[pathlib.Path, str] = {}
    for pattern in ("*.py", "*.ps1"):
        for f in app_root.rglob(pattern):
            if "migration" in f.relative_to(app_root).parts:
                continue
            try:
                per_file[f] = f.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
    corpus = "".join(per_file.values())
    init_corpus = per_file.get(app_root / "init.py", "")
    governance_generic_read = (
        "module='governance'" in init_corpus or 'module="governance"' in init_corpus)

    expectations = {r[0] for r in conn.execute(
        "SELECT expectation FROM cfg_column WHERE expectation IS NOT NULL")}
    pattern_keys = {e[len("pattern:"):] for e in expectations if e.startswith("pattern:")}
    enum_names = {e[len("enum."):] for e in expectations if e.startswith("enum.")}

    orphans: list[str] = []
    for r in conn.execute("SELECT key, module FROM cfg_setting WHERE inactive=0"):
        key, module = r[0], r[1]
        if key in pattern_keys:
            continue
        if module == "governance":
            if (governance_generic_read or f'"{key}"' in init_corpus
                    or f"'{key}'" in init_corpus):
                continue
            orphans.append(f"cfg_setting {key!r} (module 'governance' — not read by "
                           f"iba/app/init.py, the startup routine)")
            continue
        if module in _NARRATIVE_MODULES:
            # Escalation #719, 2026-08-18: all 6 module='backup' rows checked individually against
            # their live value text, not assumed -- every one documents infrastructure the app's
            # own runtime never touches (Windows Scheduled Tasks, standalone scripts under
            # scripts/*.py invoked outside the dispatcher, one historical-incident note). Same KIND
            # of content the governance branch above already exempts ("no apply-the-value
            # behaviour to grep for"), just without governance's extra "must be read by init.py"
            # compliance requirement -- there is no AI-startup-compliance angle for a backup
            # schedule fact. Confirmed no module='backup' row currently DOES pass the generic
            # check below, so this loses no real detection. Writing a fake .setting() call
            # somewhere just to silence this check would be worse than the finding it replaces.
            continue
        # ".setting(" / ".required_setting(" -- escalation (2026-08-29, the no-hardcoded-locations
        # ruling): every `.setting(key, "literal")` call site project-wide was converted to
        # `.required_setting(key)` (no silent default -- lib/cfg.py's own new method, same no-
        # fallback discipline database_path() already had), so a usage check that only recognised
        # ".setting(" went blind to all 62 of them the moment the rename landed -- found live the
        # same session, fixed in the same pass rather than left for the next validate run to
        # surface as a false "orphan" wave.
        used = any((f'"{key}"' in text or f"'{key}'" in text)
                  and (".setting(" in text or ".required_setting(" in text)
                  for text in per_file.values())
        if not used and module == "database" and _DATABASE_PATH_KEY.match(key):
            # Escalation #748, 2026-08-21: Cfg.database_path() (lib/cfg.py) is the real, live
            # consumer -- `self.setting(f"database.{name}.path")`, an f-string-composed key, so
            # the literal key text never appears in source for the same-file/same-call-site scan
            # above to find. Same class of indirection the validation.py dict-lookup case already
            # documents in this function's docstring, just via string interpolation instead of a
            # dict. Verified live both ways: (a) the exact call-site text is still present in
            # cfg.py, (b) init.py's startup path-integrity check (step 3b) genuinely calls
            # database_path() for every cfg.enum('project_database') member, applying the value,
            # not just reading and discarding it. Recurred as a false positive on every validate
            # run since escalation #727 added the real consumer without ever teaching this checker
            # the new call shape -- narrowly matched (module='database' AND the exact call-site
            # text), not a blanket module exemption, so a genuinely new orphan under module=
            # 'database' would still be caught.
            cfgpy = per_file.get(app_root / "lib" / "cfg.py", "")
            if 'self.setting(f"database.{name}.path")' in cfgpy:
                used = True
        if not used:
            orphans.append(f"cfg_setting {key!r} (key not found together with a "
                           f"cfg.setting(...) call in any one file)")

    for r in conn.execute("SELECT DISTINCT name FROM cfg_enum WHERE inactive=0"):
        name = r[0]
        if name in enum_names:
            continue
        looked_up = re.search(
            r'(\.enum\(\s*|name\s*=\s*)["\']' + re.escape(name) + r'["\']', corpus)
        if not looked_up:
            orphans.append(f"cfg_enum group {name!r} (not looked up by name at runtime — "
                           f"no cfg.enum({name!r}) or cfg_enum WHERE name={name!r} call site)")
    return orphans


# Extended 2026-07-30 per the researcher's correction ("your validations is only touching settings
# and enum and not incorporating all the other config tables") — `find_orphan_configs` above was
# the ONLY usage check in the whole app, and it only ever covered `cfg_setting`/`cfg_enum`. Every
# other table got at most a structural/referential check (does an FK point somewhere real), never
# "is this actually consumed." These three close that gap for `cfg_book_order`/`cfg_connection`/
# `cfg_candidate_rule` — the same "not referenced anywhere" concern, extended past the two tables
# it happened to start with.

def find_orphan_book_order(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """`cfg.book_order()` — the WHOLE table is read as one dict, not looked up by individual key,
    so (unlike cfg_setting/cfg_enum) there is no per-row "is this ONE book used" question — the
    question is whether the accessor itself is called anywhere (same corpus-scan convention as
    `find_orphan_configs`, `migration/` excluded for the same reason: those scripts exist to
    populate/consume the seed once, not to represent ongoing app usage). ALSO flags duplicate
    book/ordinal rows — free to check while already reading the table, the same class of internal-
    coherence gap `_validate_live`'s schema checks already catch for other tables."""
    out: list[str] = []
    used = False
    for f in app_root.rglob("*.py"):
        if "migration" in f.relative_to(app_root).parts:
            continue
        try:
            text = _code_only_text(f)
        except OSError:
            continue
        if ".book_order(" in text:
            used = True
            break
    if not used:
        out.append("cfg_book_order: cfg.book_order() is not called anywhere outside migration/ "
                  "— the whole table is unused")

    books = [r[0] for r in conn.execute("SELECT book FROM cfg_book_order WHERE inactive=0")]
    for b in sorted({b for b in books if books.count(b) > 1}):
        out.append(f"cfg_book_order: {b!r} appears more than once")
    ordinals = [r[0] for r in conn.execute("SELECT ordinal FROM cfg_book_order WHERE inactive=0")]
    for o in sorted({o for o in ordinals if ordinals.count(o) > 1}):
        out.append(f"cfg_book_order: ordinal {o} is used by more than one book")
    return out


def find_orphan_connection_keys(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """Every active `cfg_connection.key` must co-occur with a `.connection(` call in the same file
    — identical methodology to `find_orphan_configs`'s cfg_setting check, just scoped to this
    table (which that check never covered)."""
    # NOT `_code_only_text` here — that blanks OUT string-literal tokens, and the thing this check
    # needs to find (`"base_url"`) is itself a string literal. `_code_only_text` only helps when
    # searching for bare CODE SYNTAX that never legitimately appears inside a string/comment (a
    # method-call pattern with no argument, like `find_orphan_book_order`'s `.book_order(` check);
    # here it would silently blank out every real call site too — caught 2026-07-30 by re-testing
    # immediately after switching to it and seeing genuinely-used keys (`base_url` et al., real
    # calls in `stepapi.py`) suddenly flagged as unused.
    per_file: dict[pathlib.Path, str] = {}
    for f in app_root.rglob("*.py"):
        if "migration" in f.relative_to(app_root).parts:
            continue
        try:
            per_file[f] = f.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
    out = []
    for r in conn.execute("SELECT key FROM cfg_connection WHERE inactive=0"):
        key = r[0]
        used = any((f'"{key}"' in text or f"'{key}'" in text) and ".connection(" in text
                  for text in per_file.values())
        if not used:
            out.append(f"cfg_connection {key!r} — not read together with a cfg.connection(...) "
                      f"call in any one file")
    return out


# The only two steps whose code calls `.candidate_rules(...)` (`candidate.seed`/`candidate.curate`
# — verified by direct grep, 2026-07-30, not guessed). Both are currently INACTIVE (2026-07-23
# candidate-system retraction) — when both are inactive, "code calls kind X but 0 active rows back
# it" is not a live gap, it's the same already-recorded retraction every other inactive-scoped
# check already excludes (escalation #310); skip the whole first direction of the check rather than
# re-surface a fact GOVERNANCE.md §15D already covers. The reverse direction (rows nobody calls) is
# NOT step-gated — a config value with no code path to it at all is a gap regardless of whether the
# would-be caller happens to be active right now.
_CANDIDATE_RULES_CALLER_STEPS = ("candidate.seed", "candidate.curate")
_CANDIDATE_RULES_CALL_RE = re.compile(r"\.candidate_rules\(\s*[\"']([^\"']+)[\"']")


def find_orphan_candidate_rules(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """Two-directional usage check for `cfg_candidate_rule` — "orphan" means something different
    here than a single cfg_setting key, since a row is a candidate WORD, not a configurable key:
    (a) code calls the candidate_rules accessor for a kind with ZERO active rows — the call will
    silently return an empty list, which may be exactly what's happening right now and going
    unnoticed (skipped while `_CANDIDATE_RULES_CALLER_STEPS` are both inactive — see that
    constant's comment); (b) a kind WITH active rows that no code anywhere asks for — config
    nobody reads, the same direction `find_orphan_configs` already checks for cfg_setting/
    cfg_enum.

    Deliberately raw `read_text`, NOT `_code_only_text` — `_CANDIDATE_RULES_CALL_RE` needs to see
    the actual quoted argument (the kind name IS a string literal), and `_code_only_text` blanks
    string-literal tokens out; using it here would silently blank every real call site too (caught
    2026-07-30 the same way as the `cfg_connection` check above — re-tested immediately after
    switching and found real usage in `candidate.py` suddenly detected as zero). Also deliberately
    NOT spelling out the literal call-with-a-quoted-kind syntax in prose anywhere in this docstring
    (unlike the previous draft of this exact paragraph, which did, and matched its own regex
    against this file's own text) — same class of self-collision `cfgreport.py`'s own "NOTE" already
    documents for `find_utility_config_density`."""
    called_kinds: set[str] = set()
    for f in app_root.rglob("*.py"):
        if "migration" in f.relative_to(app_root).parts:
            continue
        try:
            text = f.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        called_kinds.update(_CANDIDATE_RULES_CALL_RE.findall(text))

    active_kinds = {r[0] for r in conn.execute(
        "SELECT DISTINCT kind FROM cfg_candidate_rule WHERE inactive=0")}

    out = []
    callers_inactive = all(_step_inactive(conn, s) for s in _CANDIDATE_RULES_CALLER_STEPS)
    if not callers_inactive:
        for kind in sorted(called_kinds - active_kinds):
            out.append(f"cfg_candidate_rule: code calls candidate_rules({kind!r}) but there are "
                      f"ZERO active rows of that kind — the call will silently return []")
    for kind in sorted(active_kinds - called_kinds):
        out.append(f"cfg_candidate_rule: kind {kind!r} has active rows but no code calls "
                  f"candidate_rules({kind!r}) anywhere")
    return out


def find_bad_report_csv_table_references(conn: sqlite3.Connection) -> list[str]:
    """Every `cfg_report_csv_table.table_name` must name a real table — a DATA table (`cfg_table`)
    or a `cfg_*` infrastructure table — or the wildcard `cfg_*` prefix `reportkit.write_csv_pairing`
    itself recognises (a literal trailing `*`) — **or be marked `virtual`** (a row_filter-supplied
    CSV pairing: the handler computes the rows itself and passes them to `write_csv_pairing` under
    this `table_name` as a key, so no `SELECT * FROM {table_name}` ever runs against it — see
    `write_csv_pairing`'s `row_filter` parameter). The same referential-integrity discipline
    `find_report_step_references`/`_validate_live`'s write-grant check already apply to `.step`/
    write-grant table references — never applied to THIS table's own column before (2026-07-30,
    researcher: "your validations is only touching settings and enum"). A hard structural fault:
    a CSV pairing for a table that doesn't exist AND isn't `virtual` would crash `write_csv_pairing`
    the moment that report actually runs, not just sit as a cosmetic gap.

    **`virtual` added 2026-08-15** — this check itself was a false positive against two entries
    that were never broken (`report.registry`/`word_registry_strong_pairing`,
    `report.cluster`/`strong_without_cluster`, both legitimate `row_filter` keys), re-raised across
    three separate `configmaint.validate` escalations (#591, #597, #642) over five days because the
    check had no way to represent "this is intentionally not a literal table." A `virtual` row
    without a `join_note` is still flagged — the exemption cannot silently hide an actually-wrong
    entry with no explanation attached."""
    # database='iba' -- this checks cfg_report_csv_table references against THIS app's own data
    # tables (escalation #653, 2026-08-17, widened cfg_table to also carry bible_research.db's).
    data_tables = {r[0] for r in conn.execute("SELECT name FROM cfg_table WHERE database='iba'")}
    cfg_tables = {r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'cfg\\_%' ESCAPE '\\'")}
    known = data_tables | cfg_tables
    out = []
    for r in conn.execute(
            "SELECT DISTINCT step, table_name, virtual, join_note FROM cfg_report_csv_table "
            "WHERE inactive=0"):
        name, is_virtual, join_note = r[1], r[2], r[3]
        if is_virtual:
            if join_note is None:
                out.append(f"schema: cfg_report_csv_table ({r[0]}).table_name {name!r} is marked "
                          f"virtual but carries no join_note explaining what it is")
            continue
        if name.endswith("*"):
            prefix = name[:-1]
            if not any(t.startswith(prefix) for t in known):
                out.append(f"schema: cfg_report_csv_table ({r[0]}) wildcard {name!r} matches no "
                          f"known table")
            continue
        if name not in known:
            out.append(f"schema: cfg_report_csv_table ({r[0]}).table_name {name!r} is not a "
                      f"known data or cfg_* table")
    return out


# Writer identities in cfg_write_grant that are NOT a cfg_step.step — dispatcher/API internals,
# not steps. Fallback only — the real value is `cfg_enum` group `writer_identity` (added
# 2026-07-29); kept byte-identical to that group's proposed values so this check works correctly
# even before the enum rows are approved. Found 2026-07-29: cfg_write_grant.writer was never
# checked against anything at all — a typo'd writer name would have gone unnoticed structurally
# forever, the same class of gap `find_orphan_configs` already closes for cfg_setting/cfg_enum.
_WRITER_IDENTITY_FALLBACK = ("run", "escalation", "migration",
                             "call1_meanings", "call2_getInfo", "call3_strong")


def find_report_step_references(conn: sqlite3.Connection) -> list[str]:
    """Every active `cfg_report`/`cfg_report_section`/`cfg_report_csv_table.step` must name a
    currently-ACTIVE `cfg_step.step` — the same discipline `_validate_live`'s existing `on_fail`
    check already applies to `cfg_on_fail.step`, extended to the three report-shape tables that
    were never checked against anything at all. A hard structural fault: a report row for a step
    that doesn't exist (or was retired) is broken plumbing, not a judgement call."""
    active_steps = {r[0] for r in conn.execute("SELECT step FROM cfg_step WHERE inactive=0")}
    out: list[str] = []
    for table in ("cfg_report", "cfg_report_section", "cfg_report_csv_table"):
        for r in conn.execute(f'SELECT DISTINCT step FROM "{table}" WHERE inactive=0'):
            if r[0] not in active_steps:
                out.append(f"schema: {table}.step {r[0]!r} is not a currently-active cfg_step")
    return out


def find_unknown_write_grant_writers(conn: sqlite3.Connection,
                                     writer_identities: set[str] | None = None) -> list[str]:
    """Every active `cfg_write_grant.writer` must resolve to a currently-ACTIVE `cfg_step.step`,
    OR be a declared non-step writer identity (dispatcher/API internals — see
    `_WRITER_IDENTITY_FALLBACK`/`enum.writer_identity`). Anything else is an unchecked typo/orphan
    reference — found live 2026-07-29 to be clean today, but never actually checked before."""
    active_steps = {r[0] for r in conn.execute("SELECT step FROM cfg_step WHERE inactive=0")}
    identities = writer_identities or set(_WRITER_IDENTITY_FALLBACK)
    out: list[str] = []
    # database='iba' -- escalation #680 widened cfg_write_grant to also (eventually) carry
    # bible_research.db-scoped grants; this checks THIS app's own step/writer coherence.
    for r in conn.execute("SELECT DISTINCT writer FROM cfg_write_grant WHERE database='iba' "
                          "AND inactive=0"):
        if r[0] not in active_steps and r[0] not in identities:
            out.append(f"schema: cfg_write_grant.writer {r[0]!r} is not an active cfg_step and "
                      f"not a declared writer identity")
    return out


def find_cfg_tables_missing_configmaint_grant(conn: sqlite3.Connection) -> list[str]:
    """Every real `cfg_*` RULE table must have a `cfg_write_grant` row for writer
    `configmaint.propose` — that's the one sanctioned path `USER-GUIDE.md` §9 states for changing
    any `cfg_*` row, and `governance.config_control` states it as a blanket rule with no carve-out.
    Found 2026-08-17 the hard way: `cfg_escalation`/`cfg_method_rule`/`cfg_quality_check`/`cfg_index`
    had silently shipped with NO grant at all — `cfg_method_rule`'s gap is the live, still-unfixed
    root cause of escalations #539/#550 (crashed 2026-08-06/07, reproduced identically 2026-08-17).
    A hard structural fault, not a judgement call: a `cfg_*` table nobody can write through the
    sanctioned gate is broken plumbing the same class as an unresolved `on_fail`/`write_grant`
    reference, just the opposite direction (a table with zero grants, not a grant naming a dead
    step).

    `category='rule'` (escalation #1146, 2026-08-31): scoped to genuine rule tables only, not
    every cfg_-prefixed name — `cfg_change_detail`/`cfg_change_log` are audit logs, and this check
    would otherwise immediately re-flag their now-deliberately-revoked configmaint.propose grant as
    a fresh "missing grant" finding, undoing the fix this same escalation made."""
    cfg_tables = {r[0] for r in conn.execute(
        "SELECT name FROM cfg_table WHERE database='iba' AND name LIKE 'cfg\\_%' ESCAPE '\\' "
        "AND category='rule'")}
    granted = {r[0] for r in conn.execute(
        "SELECT table_name FROM cfg_write_grant WHERE writer='configmaint.propose' "
        "AND database='iba'")}
    missing = sorted(cfg_tables - granted)
    return [f"schema: {t!r} has no cfg_write_grant row for writer 'configmaint.propose' — "
            f"nothing can maintain it through the sanctioned gate (governance.config_control)"
            for t in missing]


def find_unregistered_tables_and_columns(conn: sqlite3.Connection,
                                         project_root: pathlib.Path) -> list[str]:
    """`governance.tables`/`governance.table_columns`: every table in the project, in EITHER
    database, must be listed in `cfg_table`/`cfg_column` — "applies to all databases". Nothing
    ever actually checked this against the two databases' LIVE schema before this (found live
    2026-08-30, escalation #1058's follow-on: `finding_verse_index`, built and populated with
    475,790 rows on 2026-08-29, had zero `cfg_table`/`cfg_column` rows — a config-driven registry
    the researcher has spent six weeks building, silently out of sync with the schema it exists to
    describe, discovered by accident rather than caught by validation). This is that check, for
    both `iba` and `bible_research`, run every `configmaint.validate`.

    Four shapes of drift, all hard errors (small scale, confirmed live — single digits per
    database — not the kind of large pre-existing backlog `find_unregistered_project_scripts`
    above has to treat as advisory; unlike that check, there is no known backlog here to protect
    against drowning in, so nothing is lost by failing loudly on all four immediately):
    1. a live table with no `cfg_table` row at all;
    2. an ACTIVE (`inactive=0`) `cfg_table` row naming a table that no longer exists live (a drop/
       rename that was never reflected — should have been marked `inactive=1`, not left active
       and pointing at nothing);
    3. for every table that's both live and actively registered — a live column with no
       `cfg_column` row;
    4. an ACTIVE `cfg_column` row naming a column that no longer exists on its live table.

    `sqlite_%` tables are skipped (SQLite's own internal bookkeeping, never a project table)."""
    def _db_path(name: str) -> pathlib.Path:
        r = conn.execute("SELECT value FROM cfg_setting WHERE key=?",
                         (f"database.{name}.path",)).fetchone()
        return project_root / json.loads(r["value"])

    out: list[str] = []
    for db_key in ("iba", "bible_research"):
        live_conn = conn if db_key == "iba" else sqlite3.connect(_db_path(db_key))
        try:
            live_tables = {r[0] for r in live_conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")}
            registered = {r["name"]: r["inactive"] for r in conn.execute(
                "SELECT name, inactive FROM cfg_table WHERE database=?", (db_key,))}
            for t in sorted(live_tables - registered.keys()):
                out.append(f"schema: table {t!r} ({db_key}.db) exists live but has no cfg_table "
                          f"row (governance.tables)")
            for t in sorted(n for n, inactive in registered.items()
                            if inactive == 0 and n not in live_tables):
                out.append(f"schema: cfg_table {t!r} ({db_key}) is active but the table no longer "
                          f"exists live — dropped/renamed without updating cfg_table.inactive")
            for t in sorted(live_tables & registered.keys()):
                if registered[t] != 0:
                    continue          # table itself inactive -- its columns aren't checked either
                live_cols = {r[1] for r in live_conn.execute(f'PRAGMA table_info("{t}")')}
                reg_cols = {r["name"]: r["inactive"] for r in conn.execute(
                    "SELECT name, inactive FROM cfg_column WHERE database=? AND table_name=?",
                    (db_key, t))}
                for c in sorted(live_cols - reg_cols.keys()):
                    out.append(f"schema: {t}.{c} ({db_key}.db) exists live but has no cfg_column "
                              f"row (governance.table_columns)")
                for c in sorted(cc for cc, inactive in reg_cols.items()
                                if inactive == 0 and cc not in live_cols):
                    out.append(f"schema: cfg_column {t}.{c} ({db_key}) is active but the column no "
                              f"longer exists live on that table")
        finally:
            if db_key != "iba":
                live_conn.close()
    return out


def find_unregistered_project_scripts(conn: sqlite3.Connection,
                                      project_root: pathlib.Path) -> list[str]:
    """Phase 0 of the engine-controls migration (escalation #672, `engine-controls-migration-plan-v3
    -20260817.md`): *"any new script or routine, anywhere in the project, must be registered in
    cfg_utility... in the same unit of work it is created"* (governance.new_utility_registration_timing)
    — this is the enforcement half of that rule. Walks the WHOLE project (not just `iba/app/lib/`,
    which `find_unregistered_lib_modules` already covers by module stem) for `.py` files with no
    matching `cfg_utility.file_path` row, matched by project-root-relative path since files outside
    `iba/app/` don't share `iba/app/lib/`'s one-unique-stem-per-file guarantee. `temp_*` files are
    exempt (governance.scripts_and_routines' own carve-out for throwaway scripts).

    **ADVISORY, not a hard error** — deliberate, same reasoning as `find_stale_governance_docs`:
    at the moment this check was built, ~345 pre-existing files (`engine/`, `scripts/`, `research/`,
    `iba/prototype/`, `iba/scripts/`) are unregistered by design (that's Phase 2 of the same plan,
    not yet executed) — a hard error here would fail `configmaint.validate` wall-to-wall for a known,
    already-tracked backlog rather than catch NEW drift, which is this check's actual job. Revisit
    once Phase 2 substantially closes the backlog; hard-failing on it now would drown the signal
    this exists to give, not strengthen it."""
    # Checked against EVERY path component, not just parts[0] — found live while verifying this
    # check: scripts/analytics/venv/ is a NESTED virtualenv (3,042 site-packages .py files) that a
    # top-level-only exclusion completely misses, turning 345 real findings into 3,100+ noise.
    _EXCLUDE_ANYWHERE = {".git", "archive", "__pycache__", ".venv", "venv", "node_modules",
                        "site-packages"}
    registered = {pathlib.PurePosixPath(r[0]) for r in
                 conn.execute("SELECT file_path FROM cfg_utility")}
    out: list[str] = []
    for f in sorted(project_root.rglob("*.py")):
        rel = f.relative_to(project_root)
        parts = rel.parts
        if not parts:
            continue
        if _EXCLUDE_ANYWHERE & set(parts):
            continue
        if parts[0] == "iba" and len(parts) > 1 and parts[1] == "app":
            continue  # iba/app/ covered by find_unregistered_lib_modules + handler/tool/migration checks
        if f.stem == "__init__" or f.stem.startswith("temp_"):
            continue
        rel_posix = pathlib.PurePosixPath(rel.as_posix())
        if rel_posix not in registered:
            out.append(f"{rel_posix} has no cfg_utility row (governance.new_utility_registration_timing)")
    return out


def find_filled_by_referencing_inactive_step(conn: sqlite3.Connection) -> list[str]:
    """`cfg_column.filled_by` naming a step that is now `inactive=1` — ADVISORY, not a hard error,
    because the correct fix needs a human judgement call per column (is the column now genuinely
    dormant, as `passage.rule`/`.source` are, or has a DIFFERENT currently-active mechanism quietly
    taken over — `passage.created_at`/`verse_passage.created_at` are in fact written by
    `lib/passagetrack.py` today, not `passage.build`, and simply clearing `filled_by` there would
    be its own new inaccuracy). Found 2026-07-29: 21 columns across the candidate/passage
    retirements had this defect, silent, because nothing checked it — see
    `passage-config-full-extract-20260729.md`."""
    inactive_steps = {r[0] for r in conn.execute("SELECT step FROM cfg_step WHERE inactive=1")}
    if not inactive_steps:
        return []
    out: list[str] = []
    for r in conn.execute(
            "SELECT table_name, name, filled_by FROM cfg_column WHERE filled_by IS NOT NULL"):
        if r["filled_by"] in inactive_steps:
            out.append(f"{r['table_name']}.{r['name']} filled_by={r['filled_by']!r} "
                      f"(an inactive step) — confirm dormant or update to the real current writer")
    return out


def find_missing_report_paths(conn: sqlite3.Connection) -> list[str]:
    """Every ACTIVE quality-check step (QUALITY_CHECK_REPORT_PATH) must have its output-path
    setting actually present, non-null, and active in cfg_setting — the code-backed enforcement of
    governance.reports_must_persist. A registered quality-check step with no report path means
    its findings can only ever live in a terminal print + an escalation row, which is exactly the
    standard violation the researcher found and required fixed 2026-07-21. A retired (inactive)
    step is skipped entirely (escalation #310)."""
    missing = []
    for step, (table, key) in QUALITY_CHECK_REPORT_PATH.items():
        if _step_inactive(conn, step):
            continue
        row = conn.execute(
            f'SELECT value FROM "{table}" WHERE key=? AND inactive=0', (key,)).fetchone()
        if not row or not row[0]:
            missing.append(f"{step} has no active {key} setting in {table} — its findings would "
                          f"not persist to a report")
    return missing


def find_stale_governance_docs(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """`GOVERNANCE.md`'s own §8 rule (LIVE `cfg_setting` rows `governance.governance_md_on_rule_change`/
    `build_md_on_code_change`, escalations #238/#239): any `cfg_*` rule change updates GOVERNANCE.md,
    same unit of work — a rule §8 itself named as "follow-up work, not done in this pass" on
    2026-07-22, still unbuilt when a 2026-07-29 audit found 7 real rule changes (2026-07-26 to
    2026-07-28) with no matching GOVERNANCE.md entry. ADVISORY only (a doc update is a human act
    this can prompt, not perform): flags if the newest applied `cfg_change_detail` row is more
    recent than GOVERNANCE.md's own last-modified time. A coarse signal (a doc edit unrelated to
    config also resets the clock) — good enough to prompt a look, not precise enough to hard-fail
    on."""
    row = conn.execute("SELECT MAX(applied_at) FROM cfg_change_detail").fetchone()
    latest_change = row[0] if row else None
    if not latest_change:
        return []
    gov = app_root / "GOVERNANCE.md"
    if not gov.exists():
        return [f"{gov} does not exist — cannot check currency against the newest applied "
                f"config change ({latest_change})"]
    import datetime
    gov_mtime = datetime.datetime.fromtimestamp(
        gov.stat().st_mtime, tz=datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    if gov_mtime < latest_change:
        return [f"GOVERNANCE.md was last modified {gov_mtime}, before the newest applied "
                f"cfg_change_detail row ({latest_change}) — check whether that change needs an "
                f"entry (GOVERNANCE.md §8's own rule)"]
    return []


def find_report_version_clutter(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """`lib/reportkit.oneoff_path()` was found 2026-08-08 (BUILD.md §83) to version correctly
    (`-v2`/`-v3`/...) but never archive the superseded version — every reconciliation report and
    several extract tools accumulated their whole lineage flat in `governance.oneoff_report_dir`
    forever, unlike `write_report`'s own reports (fixed 2026-08-05, §60) which correctly hold
    exactly one live file per stem. Fixed at the source (`oneoff_path` now archives before writing)
    — this is the ACTIVE detector that a future regression (a new caller bypassing `oneoff_path`,
    or the archiving logic breaking) doesn't silently recur unnoticed, matching this app's own
    "not stated, enforced" convention rather than trusting the fix to hold by inspection alone."""
    from . import reportkit
    out_dir_setting = _cfg_setting(conn, "governance.oneoff_report_dir")
    if not out_dir_setting:
        return []
    # governance.oneoff_report_dir's value ("iba/app/reports/") is already relative to the REPO
    # ROOT (same resolution oneoff_path() itself relies on, via cwd) -- NOT relative to app_root
    # (iba/app, one level in) the way other checks in this file use it. app_root.parent.parent
    # recovers the repo root correctly either way, without assuming cwd.
    out_dir_path = pathlib.Path(out_dir_setting)
    out_dir = out_dir_path if out_dir_path.is_absolute() else app_root.parent.parent / out_dir_path
    findings = []
    for base, items in reportkit.group_oneoff_versions(out_dir).items():
        if len(items) > 1:
            versions = sorted(v for v, _ in items)
            findings.append(f"{base}: {len(items)} versions simultaneously live "
                           f"({versions}) — only the highest should be; the rest belong in "
                           f"{out_dir_setting.rstrip('/')}/archive/")
    return findings


def _cfg_setting(conn: sqlite3.Connection, key: str):
    import json
    r = conn.execute("SELECT value FROM cfg_setting WHERE key=? AND inactive=0", (key,)).fetchone()
    return json.loads(r["value"]) if r else None


def find_unregistered_lib_modules(conn: sqlite3.Connection, app_root: pathlib.Path) -> list[str]:
    """Every `iba/app/lib/*.py` module must have a `cfg_utility` row (added 2026-07-29, Phase 4 of
    `PLAN-config-system-remediation-v1-20260729.md`) — a NEW module added later and never
    registered would otherwise be invisible to `find_utility_config_density` below, the same
    "found by chance, not by a check" gap this whole registry exists to close. ADVISORY: a brand
    new file mid-edit is a normal transient state, not a coherence error."""
    registered = {r[0] for r in conn.execute("SELECT module FROM cfg_utility")}
    lib_dir = app_root / "lib"
    out = []
    if not lib_dir.exists():
        return out
    for f in sorted(lib_dir.glob("*.py")):
        if f.stem == "__init__":
            continue
        if f.stem not in registered:
            out.append(f"iba/app/lib/{f.name} has no cfg_utility row — run "
                      f"migration/bootstrap_cfg_utility.py to register it")
    return out


# Every real method the `Cfg` class exposes — computed live from the class itself (not a hand-
# maintained list) so this stays accurate as `Cfg` grows. Used to detect genuine config-consumption
# call sites (`<anything>.setting(`, `<anything>.tables(`, ...), not just the two most common ones.
# Escalation review 2026-07-30 (`cfg-utility-density-check-review-20260730.md`) found this check's
# OWN pattern was the bug for two modules: `db.py` genuinely consumes config via `.tables()`/
# `.columns()`/`.unique_key()` — real usage the old `.setting(`/`.enum(` -only pattern never
# counted; `dbsnapshot.py` genuinely calls `.setting(` but on a `Cfg` instance bound to `c`, not
# `cfg` — the old pattern hardcoded the literal text `cfg.setting(`/`cfg.enum(`, missing any other
# variable name entirely.
_STRING_TOKEN_TYPES = {tokenize.COMMENT, tokenize.STRING}
for _name in ("FSTRING_START", "FSTRING_MIDDLE", "FSTRING_END"):    # Python 3.12+ (PEP 701) only
    if hasattr(tokenize, _name):
        _STRING_TOKEN_TYPES.add(getattr(tokenize, _name))


def _code_only_text(path: pathlib.Path) -> str:
    """Source with every COMMENT/STRING(/f-string literal piece) span blanked to spaces IN PLACE —
    so a call-site scan can't be fooled by a docstring or comment that merely MENTIONS the pattern
    in prose, while every real call site keeps its EXACT original adjacency (`cfg.setting(` stays
    literally `cfg.setting(`, not spread apart). Found 2026-07-30, twice in one session:
    `cfgreport.py` (fixed by rewording one line) and then `cfgquality.py` itself — THIS file's own
    docstrings talk about `.setting(`/`.enum(`/`.tables(` (they document the exact check below),
    which falsely satisfied the old raw-text substring scan. First attempt at this fix (joining
    non-string tokens with a space) was itself wrong — it broke every REAL match too, by inserting
    spaces between `cfg`/`.`/`setting`/`(`, caught immediately by re-testing `db.py`/`dbsnapshot.py`
    (known-good cases) and finding them suddenly failing. Blanking spans in the original string,
    same length, fixes both directions at once. Falls back to raw text if a file doesn't tokenize
    cleanly (better a possible false negative than a hard crash on some edge-case file)."""
    text = path.read_text(encoding="utf-8", errors="ignore")
    try:
        tokens = list(tokenize.generate_tokens(io.StringIO(text).readline))
    except (tokenize.TokenError, IndentationError, SyntaxError, ValueError):
        return text
    line_starts = [0]
    for line in text.splitlines(keepends=True):
        line_starts.append(line_starts[-1] + len(line))
    chars = list(text)
    for tok in tokens:
        if tok.type not in _STRING_TOKEN_TYPES:
            continue
        start = line_starts[tok.start[0] - 1] + tok.start[1]
        end = line_starts[tok.end[0] - 1] + tok.end[1]
        for i in range(start, min(end, len(chars))):
            if chars[i] != "\n":
                chars[i] = " "
    return "".join(chars)


def _cfg_method_pattern() -> re.Pattern:
    from .cfg import Cfg
    methods = [m for m in dir(Cfg) if not m.startswith("_") and callable(getattr(Cfg, m))
              and m != "close"]
    return re.compile(r"\.(?:" + "|".join(re.escape(m) for m in methods) + r")\(")


_CFG_METHOD_RE = None  # lazy: built on first use, not at import time (avoids a circular import —
                       # cfg.py doesn't import cfgquality, but this keeps the dependency one-way)


def find_utility_config_density(conn: sqlite3.Connection) -> list[str]:
    """Every ACTIVE, NON-EXEMPT `cfg_utility` module with ZERO real `Cfg`-method call sites in its
    own file (any method `Cfg` exposes — `.setting(`/`.enum(` and every schema/write-grant/sequence
    method alike, under any variable name — see `_cfg_method_pattern`) — ADVISORY, the same "could
    be legitimate" caveat `find_orphan_configs` already carries: a module like `retention.py`/
    `seedreport.py` genuinely has no config of its own because its caller resolves paths for it (a
    legitimate zero — now declared via `cfg_utility.config_exempt`, not re-derived every run); a
    module like `lexiconparse.py` — six regexes and a hardcoded tag-set deciding a real parse, zero
    `Cfg` reference anywhere — is the actual gap this check exists to surface (found by hand
    2026-07-29, `core-module-config-intent-vs-effect-20260729.md`). `cfg_utility.file_path` is
    already repo-root-relative (e.g. `iba/app/lib/cfg.py`) — resolved against the CWD, which every
    PS entry point already sets to the repo root (`Set-Location $RepoRoot`), not recombined with
    any other path."""
    global _CFG_METHOD_RE
    if _CFG_METHOD_RE is None:
        _CFG_METHOD_RE = _cfg_method_pattern()
    out = []
    for r in conn.execute(
            "SELECT module, file_path FROM cfg_utility WHERE inactive=0 AND config_exempt=0"):
        path = pathlib.Path(r["file_path"])
        if not path.exists():
            out.append(f"cfg_utility {r['module']!r} — {path} no longer exists on disk")
            continue
        if not _CFG_METHOD_RE.search(_code_only_text(path)):
            out.append(f"cfg_utility {r['module']!r} ({path}) has zero Cfg-method call sites "
                      f"(.setting()/.enum()/.tables()/... under any variable name) — confirm this "
                      f"is a legitimate zero (mark `cfg_utility.config_exempt=1` via "
                      f"`configmaint.propose`) or a real completeness gap")
    return out


# D28 (register v9): Escalation.ps1's `[ValidateSet(...)]` literals are a THIRD, hardcoded copy of
# the escalation vocabulary — alongside the two live cfg_enum groups (escalation_next_action_manual/
# _dispatcher) lib/escalation.py actually validates against. Nothing keeps the PS copy in sync if the
# enum changes (D27's own fix wouldn't reach it without a manual edit). A drift-detection check, not
# a dynamically-querying ValidateSet — disproportionate machinery for a rarely-changing list (per the
# register's own reasoning). Maps PS -Parameter name -> (cfg_enum group, value-transform) — only the
# groups meant to be an EXACT match; -State's ValidateSet is a deliberate curated SUBSET (only the
# explicitly-settable states — raised/re-assigned/completed are system-derived) and is excluded here
# for that reason, not because it was overlooked.
_PS_VALIDATESET_ENUM_MAP = {
    "NextAction": ("escalation_next_action_manual", lambda v: v),
    "Decision": ("escalation_next_action_dispatcher", lambda v: v.lower()),
    "Type": ("escalation_type", lambda v: v),
    "AnsweredBy": ("escalation_assignee", lambda v: v),
    "AssignedTo": ("escalation_assignee", lambda v: v),
}
_VALIDATESET_RE = re.compile(
    r"\[ValidateSet\((?P<items>(?:'[^']*'|\"[^\"]*\"|\s*,\s*)+)\)\]\s*\[\w+(?:\[\])?\]\s*\$(?P<param>\w+)")
_VALIDATESET_ITEM_RE = re.compile(r"'([^']*)'|\"([^\"]*)\"")


def _ps_validateset_drift(conn: sqlite3.Connection, ps_path: pathlib.Path, ps_label: str,
                          enum_map: dict) -> list[str]:
    """Generic core: every `enum_map` parameter's `[ValidateSet(...)]` values in `ps_path` must
    exactly match the live `cfg_enum` group it's supposed to mirror. Escalation #977, 2026-08-27
    (D28/register v9) originated this for `Escalation.ps1` specifically; generalised 2026-08-28
    (escalation #971/#977) so a second PS script's own hardcoded `ValidateSet` vocabulary — e.g.
    `FolderPurpose.ps1`'s `-Type`/`-Status` — gets the same standing drift check instead of being
    a one-off, silently-driftable copy nothing ever re-verifies against its `cfg_enum` source."""
    if not ps_path.exists():
        return [f"{ps_label} not found at {ps_path} — cannot check ValidateSet drift"]
    text = ps_path.read_text(encoding="utf-8", errors="ignore")
    found: dict[str, list[str]] = {}
    for m in _VALIDATESET_RE.finditer(text):
        param = m.group("param")
        if param not in enum_map:
            continue
        items = [a or b for a, b in _VALIDATESET_ITEM_RE.findall(m.group("items"))]
        found[param] = items

    out = []
    for param, (enum_name, transform) in enum_map.items():
        if param not in found:
            out.append(f"{ps_label}: -{param} has no [ValidateSet(...)] found (expected to "
                      f"mirror cfg_enum {enum_name!r})")
            continue
        ps_values = {transform(v) for v in found[param]}
        live_values = {r[0] for r in conn.execute(
            "SELECT value FROM cfg_enum WHERE name=? AND inactive=0", (enum_name,))}
        missing = live_values - ps_values
        extra = ps_values - live_values
        if missing:
            out.append(f"{ps_label}: -{param} ValidateSet is missing {sorted(missing)} "
                      f"(present in live cfg_enum {enum_name!r} but not the PS ValidateSet)")
        if extra:
            out.append(f"{ps_label}: -{param} ValidateSet has {sorted(extra)} which is not in "
                      f"live cfg_enum {enum_name!r} — stale or renamed")
    return out


def find_escalation_ps_validateset_drift(conn: sqlite3.Connection, app_root: pathlib.Path
                                         ) -> list[str]:
    """Every `_PS_VALIDATESET_ENUM_MAP` parameter's `[ValidateSet(...)]` values must exactly match
    the live `cfg_enum` group it's supposed to mirror (after `-Decision`'s lowercase transform).
    ADVISORY — a real drift is a genuine finding needing a look, not necessarily a hard structural
    fault (the PS script may simply not have been updated yet)."""
    return _ps_validateset_drift(conn, app_root / "ps" / "Escalation.ps1", "Escalation.ps1",
                                 _PS_VALIDATESET_ENUM_MAP)


_FOLDERPURPOSE_PS_VALIDATESET_ENUM_MAP = {
    "Type": ("folder_purpose_type", lambda v: v),
    "Status": ("folder_purpose_status", lambda v: v),
}


def find_folderpurpose_ps_validateset_drift(conn: sqlite3.Connection, app_root: pathlib.Path
                                            ) -> list[str]:
    """`FolderPurpose.ps1`'s `-Type`/`-Status` `[ValidateSet(...)]` values against the live
    `cfg_enum` groups they mirror — escalation #977's own resolution: the values ARE registered in
    `cfg_enum` and `set_purpose()` validates against it live (not a hardcoded Python set), but nothing
    previously checked whether the OTHER two hardcoded copies of the same vocabulary — this PS
    script's `ValidateSet`, and `lib/folderpurpose.py:_assess_type()`/`_assess_status()`'s literal
    `return` values (Method D — necessarily hardcoded, since they're the RULES deciding which
    enum value applies, not a re-statement of what values exist) — had drifted. This check covers
    the PS copy; `auto_assess()` now validates its own literal vocabulary against the live enum at
    the top of every run (raises loudly on drift, not a silent bad write) — found live while
    resolving #977 that it hadn't been, unlike `set_purpose()`, which always had."""
    return _ps_validateset_drift(conn, app_root / "ps" / "FolderPurpose.ps1", "FolderPurpose.ps1",
                                 _FOLDERPURPOSE_PS_VALIDATESET_ENUM_MAP)


# ── PS-script vs. Excel-worksheet drift (escalation #1007 follow-on, 2026-08-29) ────────────────
# The researcher keeps two hand/generated Excel workbooks as the CLI interface to iba/app/ps:
# `iba/docs/ps tools worksheet.xlsx` (one tab per script, row 4 = flag headers, mechanically built
# from each script's own `param()` block — escalation #1004) and `iba/docs/escalation actions
# worksheet.xlsx` (the researcher's own hand-built model, one sheet, fixed action-shapes for
# Escalation.ps1 specifically). Nothing previously re-checked either workbook against the scripts
# they describe — the same class of drift the ValidateSet checks above catch for a narrower case
# (one parameter's allowed VALUES), generalised here to a script's whole PARAMETER LIST against its
# worksheet tab. Two checks, not one, because the two workbooks are structurally different: the
# generic one is a strict per-script tab match; the hand-built escalation one only supports a
# subset check (its columns are the union of several fixed action-shapes, not one row per param).
_PS_WORKSHEET_SKIP_STEMS = {"Escalation"}  # its tab is a pointer to the model sheet, not real headers
# PowerShell automatic variables that can legally appear INSIDE a param() block without being a
# parameter — e.g. `[Parameter(Mandatory = $true)]` — found live: Behaviour.ps1 and 20+ others use
# exactly this attribute, and a bare `\$(\w+)` capture can't tell "a parameter's own $Name" from
# "a literal referenced inside someone else's attribute" any other way without a real PS parser
# (disproportionate for this check's purpose, per the same reasoning _ps_validateset_drift's own
# regex approach already accepted).
_PS_AUTOMATIC_VARS = {"true", "false", "null"}


def _ps_param_names(ps_path: pathlib.Path) -> set[str]:
    """Every declared parameter name in `ps_path`'s outer `param(...)` block — bracket-depth aware
    (a naive regex closing on the first `)` would mis-close on a `[ValidateSet('a','b')]`'s own,
    earlier, closing paren) rather than a fixed-form regex like `_VALIDATESET_RE` above, which only
    has to find named ValidateSet blocks, not the whole param list."""
    text = ps_path.read_text(encoding="utf-8", errors="ignore")
    m = re.search(r"\bparam\s*\(", text, re.I)
    if not m:
        return set()
    i, depth = m.end(), 1
    while i < len(text) and depth > 0:
        if text[i] == "(":
            depth += 1
        elif text[i] == ")":
            depth -= 1
        i += 1
    body = text[m.end():i - 1]
    return {mm.group(1) for mm in re.finditer(r"\$(\w+)", body)
           if mm.group(1).lower() not in _PS_AUTOMATIC_VARS}


def _load_worksheet_setting(conn: sqlite3.Connection, key: str, project_root: pathlib.Path
                            ) -> tuple[pathlib.Path | None, str | None]:
    """Resolves a `governance.*_worksheet_path` setting to a real file — returns (path, None) or
    (None, error-string) so callers can surface a config gap as a finding rather than crashing."""
    row = conn.execute(
        "SELECT value FROM cfg_setting WHERE key=? AND inactive=0", (key,)).fetchone()
    if not row:
        return None, f"{key} is not set — cannot check PS/worksheet drift"
    wb_path = project_root / json.loads(row[0])
    if not wb_path.exists():
        return None, f"{key} points at {wb_path} — file does not exist"
    return wb_path, None


def _index_ps_worksheet_tabs(wb) -> dict[str, "openpyxl.worksheet.worksheet.Worksheet"]:
    """Maps a script's own embedded path cell (e.g. `iba\\app\\ps\\Behaviour.ps1`, found live in
    row 6 of its tab, col A — each tab's compiled-command formula references it) back to that
    worksheet, scanning every tab once. Matching on the embedded path rather than the tab NAME
    because Excel's 31-char sheet-name limit already forces some tabs to a shortened name (e.g.
    `create-passages-by-book-view-and-export.ps1` → tab `passages-by-book-view-export`) — the
    embedded path is never truncated and stays the one reliable anchor."""
    idx: dict[str, object] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=3, values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.strip().lower().endswith(".ps1"):
                    idx[cell.strip().replace("/", "\\").lower()] = ws
    return idx


def find_ps_worksheet_drift(conn: sqlite3.Connection, app_root: pathlib.Path,
                            project_root: pathlib.Path) -> list[str]:
    """Every `iba/app/ps/*.ps1` script's live `param()` names against its own tab's row-4 flag
    headers in `governance.ps_worksheet_path` (`iba/docs/ps tools worksheet.xlsx`) — a parameter
    added/removed/renamed on the script with no matching tab update is exactly the drift the
    researcher asked this check to catch (2026-08-29: "any change to any PS instruction will find
    its way into the two excel worksheets"). ADVISORY — a real finding needing a look, not
    necessarily wrong (the tab may just not have been updated yet). `Escalation.ps1` is skipped —
    its tab is a deliberate pointer to the OTHER worksheet, checked separately below."""
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl not importable — cannot check PS/worksheet drift"]
    wb_path, error = _load_worksheet_setting(conn, "governance.ps_worksheet_path", project_root)
    if error:
        return [error]
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    tabs = _index_ps_worksheet_tabs(wb)
    out = []
    for ps_file in sorted((app_root / "ps").glob("*.ps1")):
        if ps_file.stem in _PS_WORKSHEET_SKIP_STEMS:
            continue
        key = f"iba\\app\\ps\\{ps_file.name}".lower()
        ws = tabs.get(key)
        if ws is None:
            out.append(f"{ps_file.name}: no tab found in {wb_path.name} (expected an embedded "
                      f"path cell {key!r} in some tab's first rows)")
            continue
        live = _ps_param_names(ps_file)
        sheet = {c[1:] for row in ws.iter_rows(min_row=4, max_row=4, values_only=True)
                for c in row if isinstance(c, str) and c.startswith("-")}
        # PowerShell parameter names are case-insensitive; compare on that basis (found live:
        # Escalation.ps1's own '-Action' vs. the model sheet's '-action' would false-positive
        # otherwise — same fix applied to both checks here for consistency).
        sheet_lower = {s.lower() for s in sheet}
        live_lower = {p.lower() for p in live}
        missing = {p for p in live if p.lower() not in sheet_lower}
        extra = {s for s in sheet if s.lower() not in live_lower}
        if missing:
            out.append(f"{ps_file.name}: {wb_path.name} tab {ws.title!r} is missing flag "
                      f"column(s) {sorted(missing)} — the script has these parameters now")
        if extra:
            out.append(f"{ps_file.name}: {wb_path.name} tab {ws.title!r} has flag column(s) "
                      f"{sorted(extra)} the script no longer has — stale or renamed")
    return out


def find_escalation_worksheet_drift(conn: sqlite3.Connection, app_root: pathlib.Path,
                                    project_root: pathlib.Path) -> list[str]:
    """`Escalation.ps1`'s live `param()` names must each appear SOMEWHERE as a `-Flag` header in
    the researcher's own `governance.escalation_worksheet_path` (`iba/docs/escalation actions
    worksheet.xlsx`) — a subset check, not a per-tab exact match like `find_ps_worksheet_drift`
    above, because that workbook's one sheet is hand-built around several fixed action-shapes
    (Raise/Update/AnswerRun/...) sharing columns, not one row per parameter. Still catches the
    thing the researcher actually asked for: a new/renamed Escalation.ps1 parameter that the
    model sheet was never updated to include."""
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl not importable — cannot check PS/worksheet drift"]
    wb_path, error = _load_worksheet_setting(conn, "governance.escalation_worksheet_path",
                                             project_root)
    if error:
        return [error]
    live = _ps_param_names(app_root / "ps" / "Escalation.ps1")
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    used = {c[1:] for ws in wb.worksheets for row in ws.iter_rows(values_only=True)
           for c in row if isinstance(c, str) and c.startswith("-") and " " not in c}
    used_lower = {u.lower() for u in used}
    missing = {p for p in live if p.lower() not in used_lower}
    if missing:
        return [f"Escalation.ps1 has parameter(s) {sorted(missing)} not used as a -Flag header "
                f"anywhere in {wb_path.name} — the researcher's model sheet may need updating"]
    return []


def find_unclassified_active_steps(conn: sqlite3.Connection) -> list[str]:
    """Every ACTIVE `cfg_step` must have `kind` set (`operations` | `utility` — the researcher's
    own classification, 2026-07-30, `migration/bootstrap_step_kind.py`). A HARD error, not a
    judgement call: `run.py`'s dispatch gate (escalation, same day) refuses to dispatch a step with
    no classification at all — this is the structural check that keeps that gate's premise true
    (a NEW step, added later without a `kind`, would otherwise silently sit undispatchable with no
    coherence check pointing at why). `find_enum_violations` (value-quality) already catches an
    INVALID `kind` value; this catches a MISSING one, which that check explicitly does not (it
    skips `NULL` by design)."""
    return [f"cfg_step ({r['work_package']}, {r['step']}) is active but has no kind "
           f"(operations|utility) — classify it via configmaint.propose before it can be "
           f"dispatched (run.py now refuses undispatched steps with no kind)"
           for r in conn.execute(
               "SELECT work_package, step FROM cfg_step WHERE inactive=0 AND kind IS NULL")]


def find_unresolvable_location_settings(conn: sqlite3.Connection, project_root: pathlib.Path
                                       ) -> list[str]:
    """Every location-shaped config value (`cfg_setting` AND every per-module table shaped like it
    — `cfg_prose`, `cfg_passage`, ...; see `lib/folderpurpose.location_settings()`, the one shared
    enumeration both this check and `folderpurpose`'s Method B use) must resolve to a real folder
    on disk, project-root-relative. ADVISORY, not a hard structural fault — a setting can point
    ahead of a folder not yet created by design (a report path whose folder gets `mkdir(parents=
    True)`'d on first write), so this flags for a look rather than failing the whole validate run.

    Added 2026-08-28 (researcher, escalation #971/#976): "configmaint should validate every
    location reference in every config as part of its validation routine." Direct cause: `cfg_prose.
    prose.edit_file_dir` pointed at `outputs/markdown/prose-edits`, a folder physically moved away
    (to `Workflow/Programme/prose-edits`) during the 2026-08-27 folder reorg with the setting never
    updated to follow — undetected until the researcher noticed a DIFFERENT gap (`folder_purpose`
    missing a `governed_by_setting`) and asked why. This check makes that class of drift a standing,
    automatic finding instead of something found by chance."""
    # Local import: folderpurpose lives in lib/ alongside this module; importing at call time (not
    # module level) avoids a hard dependency for every cfgquality caller that never needs it.
    from . import folderpurpose as fp_mod

    out = []
    for table, key, raw_value in fp_mod.location_settings(conn):
        norm = fp_mod.normalize_setting_value(raw_value)
        if norm is None:
            continue  # not a plausible bare folder path (a sentence, a {template}, a JSON list)
        if not (project_root / norm).is_dir():
            out.append(f"{table}.{key} = {raw_value!r} — {norm!r} does not exist as a folder "
                      f"on disk (project-root-relative)")
    return out


# `filing` behaviour class item 4 (escalation #863/#971/#992): a write that hand-imitates the
# same-day -v{n} versioning pattern instead of calling filingkit.versioned_path()/reportkit.
# oneoff_path(). Matches an f-string/format literal building "...-v{...}..." or "...-v" + str(...)
# by hand — the two ways this codebase's own existing writers were found (2026-08-28 survey) to do
# it before the shared utility existed.
_HAND_VERSION_RE = re.compile(r"-v\{[^}]*\}|-v[\"']\s*\+")
_FILINGKIT_CALL_RE = re.compile(r"\b(?:versioned_path|oneoff_path)\s*\(")
_FILING_SCAN_EXCLUDE_ANYWHERE = {".git", "archive", "__pycache__", ".venv", "venv",
                                "node_modules", "site-packages", "migration",
                                # 2026-08-28: `engine/` is superseded (CLAUDE.md top banner,
                                # 2026-08-15) and its own `-v{n}` matches turned out to be a
                                # reference table of NAMING PATTERNS (documentary string data),
                                # not real file-writing code — genuinely not this check's target.
                                "engine"}


def _strip_comments(text: str) -> str:
    """Blanks Python COMMENT token spans (only — string literals stay live) so a comment merely
    MENTIONING '-v{n}' in prose (found live: escalation.py's own comment describing that its
    caller ALREADY goes through reportkit.write_report()) doesn't false-positive the same way
    `cfgquality._code_only_text`'s docstring problem does for call-site scans, just inverted."""
    try:
        tokens = list(tokenize.generate_tokens(io.StringIO(text).readline))
    except (tokenize.TokenError, IndentationError, SyntaxError, ValueError):
        return text
    line_starts = [0]
    for line in text.splitlines(keepends=True):
        line_starts.append(line_starts[-1] + len(line))
    chars = list(text)
    for tok in tokens:
        if tok.type != tokenize.COMMENT:
            continue
        start = line_starts[tok.start[0] - 1] + tok.start[1]
        end = line_starts[tok.end[0] - 1] + tok.end[1]
        for i in range(start, min(end, len(chars))):
            if chars[i] != "\n":
                chars[i] = " "
    return "".join(chars)


def find_hand_rolled_versioning(conn: sqlite3.Connection, project_root: pathlib.Path
                                ) -> list[str]:
    """`filing` behaviour class rule 4 (`cfg_behaviour_rule` 'archiving-trigger', plus the
    naming-shape rule) as a standing check, not just a stated rule nobody re-verifies: every ACTIVE
    (`cfg_utility.inactive=0`, or unregistered — same scope `pathaudit.py` uses) `.py` file that
    builds a `-v{n}`-shaped filename by hand should be calling `filingkit.versioned_path()` (or its
    `reportkit.oneoff_path()` wrapper) instead of reimplementing the same-day-bump/archive-before-
    write logic itself. ADVISORY — a real hit needs a look (migrate the call site), not necessarily
    a hard fault; `iba/app/migration/` excluded for the same reason `pathaudit.py` excludes it (a
    migration's own filename literals are seed data, not a report-writing call site)."""
    inactive = {r[0] for r in conn.execute("SELECT file_path FROM cfg_utility WHERE inactive=1")}
    out = []
    for f in sorted(project_root.rglob("*.py")):
        rel = f.relative_to(project_root)
        parts = rel.parts
        if not parts or _FILING_SCAN_EXCLUDE_ANYWHERE & set(parts):
            continue
        if f.stem == "__init__" or f.stem.startswith("temp_"):
            continue
        rel_posix = rel.as_posix()
        if rel_posix in inactive:
            continue
        if rel_posix in ("iba/app/lib/filingkit.py", "iba/app/lib/reportkit.py"):
            continue  # the canonical implementation itself, not a caller reinventing it
        try:
            text = f.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        code_text = _strip_comments(text)
        if _HAND_VERSION_RE.search(code_text) and not _FILINGKIT_CALL_RE.search(text):
            out.append(f"{rel_posix} builds a -v{{n}} filename by hand — no filingkit."
                      f"versioned_path()/reportkit.oneoff_path() call site in the same file")
    return out
